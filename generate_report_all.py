# ============================================================
# generate_report_all.py -- ANTIGRAVITI Multi-Version Excel Reporter
#
# AMAC:
#   V1, V2 ve V2.1 icin her birinin kendi ayri Excel dosyasini
#   olusturur. Islem gecmisi, acik pozisyonlar, ozet istatistikler
#   ve sinyal kayitlari sekmelenmiş sayfalarda goruntulenir.
#
# KULLANIM:
#   python generate_report_all.py
#   python generate_report_all.py --output-dir raporlar/
# ============================================================

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    print("HATA: openpyxl yuklu degil. Lutfen yukleyin: pip install openpyxl")
    sys.exit(1)

# Windows stdout encoding fix
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

PROJECT_ROOT = Path(__file__).parent

# ---- Versiyon Tanimi ----
VERSIONS = [
    {
        "label":    "V1",
        "title":    "Strateji V1 - Long-Only Klasik",
        "log_dir":  PROJECT_ROOT / "logs_v1",
        "color":    "0070C0",   # Mavi
        "accent":   "BDD7EE",
    },
    {
        "label":    "V2",
        "title":    "Strateji V2 - Long-Only + Filtreler",
        "log_dir":  PROJECT_ROOT / "logs",
        "color":    "375623",   # Koyu yesil
        "accent":   "E2EFDA",
    },
    {
        "label":    "V2.1",
        "title":    "Strateji V2.1 - Long + SHORT",
        "log_dir":  PROJECT_ROOT / "logs_v21",
        "color":    "7030A0",   # Mor
        "accent":   "EAD1DC",
    },
]

# ---- Yardimci Fonksiyonlar ----

def load_portfolio(log_dir: Path) -> Optional[Dict]:
    path = log_dir / "portfolio_state.json"
    if not path.exists():
        return None
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"  UYARI: {path} yuklenemedi: {e}")
        return None


def load_signals(log_dir: Path) -> List[Dict]:
    """signals.jsonl veya signals_*.jsonl dosyalarindan sinyal kayitlarini yukler."""
    records = []
    # Tek dosya (eski format)
    single = log_dir / "signals.jsonl"
    if single.exists():
        records.extend(_read_jsonl(single))
    # Tarihli dosyalar
    for f in sorted(log_dir.glob("signals_*.jsonl")):
        records.extend(_read_jsonl(f))
    return records


def _read_jsonl(path: Path) -> List[Dict]:
    rows = []
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        rows.append(json.loads(line))
                    except Exception:
                        pass
    except Exception:
        pass
    return rows


def fmt_dt(iso: Optional[str]) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(iso)


def make_border(thin=True) -> Border:
    s = Side(style="thin" if thin else "medium")
    return Border(left=s, right=s, top=s, bottom=s)


# ---- Excel Stillemesi ----

def header_style(ws, row: int, cols: List[str], color: str) -> None:
    """Baslik satiri icin dolgu ve yazi stili uygular."""
    fill = PatternFill(fill_type="solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = label
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()


def data_row_style(ws, row: int, ncols: int, accent: str, is_even: bool) -> None:
    """Veri satiri renklendirmesi."""
    bg = accent if is_even else "FFFFFF"
    fill = PatternFill(fill_type="solid", fgColor=bg)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            cell.value = ""
        cell.fill = fill
        cell.border = make_border()
        cell.alignment = Alignment(vertical="center")


def auto_col_width(ws) -> None:
    """Sutun genisliklerini otomatik ayarla."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


# ---- Sayfa Olusturucular ----

def add_summary_sheet(wb: Workbook, ver: dict, state: Optional[Dict]) -> None:
    """Ozet sayfa: Bakiye, toplam PnL, win rate vb."""
    ws = wb.active
    ws.title = "Ozet"

    color  = ver["color"]
    accent = ver["accent"]

    # Baslik
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = ver["title"].upper()
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor=color)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Rapor Zamani: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ts_cell.font = Font(italic=True, size=9, color="666666")
    ts_cell.alignment = Alignment(horizontal="right")

    # Veri yoksa
    if state is None:
        ws["A4"] = "UYARI: Bu versiyon icin portfolyo verisi bulunamadi."
        ws["A4"].font = Font(color="FF0000", bold=True)
        return

    orders = state.get("orders", [])
    positions = state.get("open_positions", {})
    balance = state.get("usdt_balance", 0.0)

    closed = [o for o in orders if o.get("close_reason") and o.get("status") == "filled"]
    wins   = [o for o in closed if o.get("pnl_usdt", 0) >= 0]
    losses = [o for o in closed if o.get("pnl_usdt", 0) < 0]

    total_pnl      = sum(o.get("pnl_usdt", 0) for o in closed)
    total_fees     = sum(o.get("fee_usdt", 0) for o in orders if o.get("status") == "filled")
    gross_profit   = sum(o.get("pnl_usdt", 0) for o in wins)
    gross_loss     = sum(o.get("pnl_usdt", 0) for o in losses)
    win_rate       = len(wins) / len(closed) * 100 if closed else 0
    avg_win        = gross_profit / len(wins) if wins else 0
    avg_loss       = abs(gross_loss) / len(losses) if losses else 0
    profit_factor  = gross_profit / abs(gross_loss) if gross_loss != 0 else 0

    open_pnl  = sum(p.get("pnl_usdt", 0) for p in positions.values())
    open_cost = sum(p.get("cost_usdt", 0) for p in positions.values())
    total_val = balance + open_cost + open_pnl

    metrics = [
        ("BAKIYE VE DEGER", None, None),
        ("USDT Bakiye (serbest)", f"${balance:,.2f}", None),
        ("Acik Pozisyon Degeri", f"${open_cost:,.2f}", None),
        ("Acik Poz. PnL", f"${open_pnl:+,.2f}", "green" if open_pnl >= 0 else "red"),
        ("Toplam Portfolyo Degeri", f"${total_val:,.2f}", None),
        ("", None, None),
        ("ISLEM ISTATISTIKLERI", None, None),
        ("Toplam Kapali Islem", str(len(closed)), None),
        ("Kazanc Sayisi", str(len(wins)), "green"),
        ("Kayip Sayisi", str(len(losses)), "red"),
        ("Basari Orani", f"%{win_rate:.1f}", "green" if win_rate >= 50 else "red"),
        ("Toplam Net PnL", f"${total_pnl:+,.2f}", "green" if total_pnl >= 0 else "red"),
        ("Toplam Odenen Komisyon", f"${total_fees:.2f}", None),
        ("Brut Kar (kazanan islemler)", f"${gross_profit:+,.2f}", "green"),
        ("Brut Zarar (kaybeden islemler)", f"${gross_loss:+,.2f}", "red"),
        ("Ortalama Kazanc / Islem", f"${avg_win:,.2f}", None),
        ("Ortalama Kayip / Islem", f"${avg_loss:,.2f}", None),
        ("Kar Faktoru (Profit Factor)", f"{profit_factor:.2f}", "green" if profit_factor >= 1 else "red"),
        ("", None, None),
        ("ACIK POZISYONLAR", None, None),
        ("Acik Pozisyon Sayisi", str(len(positions)), None),
    ]

    row = 4
    for label, value, color_hint in metrics:
        if value is None and label:
            # Bolum basligi
            ws.merge_cells(f"A{row}:E{row}")
            cell = ws[f"A{row}"]
            cell.value = label
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor=ver["color"])
            cell.alignment = Alignment(horizontal="left", indent=1)
            ws.row_dimensions[row].height = 18
        elif label == "":
            ws.row_dimensions[row].height = 8
        else:
            lbl_cell = ws[f"A{row}"]
            val_cell = ws[f"B{row}"]
            lbl_cell.value = label
            val_cell.value = value
            lbl_cell.font = Font(size=10)
            lbl_cell.fill = PatternFill(fill_type="solid", fgColor=accent)
            val_cell.font = Font(bold=True, size=10,
                                  color=("375623" if color_hint == "green"
                                         else ("CC0000" if color_hint == "red" else "000000")))
            val_cell.alignment = Alignment(horizontal="right")
            lbl_cell.border = make_border()
            val_cell.border = make_border()
            ws.row_dimensions[row].height = 16

        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20


def add_orders_sheet(wb: Workbook, ver: dict, state: Optional[Dict]) -> None:
    """Tum emirler (hem acik hem kapali) sayfasi."""
    ws = wb.create_sheet("Emirler")
    color  = ver["color"]
    accent = ver["accent"]

    headers = [
        "Emir ID", "Sembol", "Yon", "Tip", "Durum",
        "Giris Fiyati", "Miktar", "Maliyet (USDT)", "Komisyon",
        "SL", "TP", "Kapanis Nedeni", "PnL (USDT)", "PnL (%)",
        "Tarih"
    ]
    header_style(ws, 1, headers, color)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    if state is None:
        ws["A2"] = "Veri bulunamadi."
        return

    orders = state.get("orders", [])
    # Tarihe gore sirala (en yeni uste)
    try:
        orders = sorted(orders, key=lambda o: o.get("timestamp", ""), reverse=True)
    except Exception:
        pass

    for i, o in enumerate(orders, 2):
        side   = o.get("side", "")
        status = o.get("status", "")
        pnl_u  = o.get("pnl_usdt", 0.0)
        pnl_p  = o.get("pnl_pct", 0.0)

        row_data = [
            o.get("order_id", ""),
            o.get("symbol", ""),
            side.upper(),
            o.get("order_type", ""),
            status.upper(),
            o.get("price", 0.0),
            o.get("amount", 0.0),
            round(o.get("price", 0) * o.get("amount", 0), 2),
            o.get("fee_usdt", 0.0),
            o.get("stop_loss", 0.0),
            o.get("take_profit", 0.0),
            o.get("close_reason", ""),
            round(pnl_u, 2),
            round(pnl_p * 100, 2),
            fmt_dt(o.get("timestamp")),
        ]

        is_even = (i % 2 == 0)
        data_row_style(ws, i, len(headers), accent, is_even)

        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            # Renk kodlama
            if j == 3:  # Yon
                cell.font = Font(bold=True,
                                  color="005500" if side == "buy" else "880000")
            if j == 13:  # PnL
                cell.font = Font(bold=True,
                                  color="005500" if pnl_u >= 0 else "880000")
            if j == 5:  # Durum
                status_colors = {"FILLED": "005500", "FAILED": "880000", "PENDING": "CC6600"}
                cell.font = Font(color=status_colors.get(status.upper(), "000000"))

    auto_col_width(ws)


def add_positions_sheet(wb: Workbook, ver: dict, state: Optional[Dict]) -> None:
    """Acik pozisyonlar sayfasi."""
    ws = wb.create_sheet("Acik Pozisyonlar")
    color  = ver["color"]
    accent = ver["accent"]

    headers = [
        "Sembol", "Yon", "Giris Fiyati", "Guncel Fiyat",
        "Miktar", "Maliyet (USDT)", "SL", "TP",
        "Anlık PnL (USDT)", "Anlık PnL (%)", "Acilis Tarihi",
        "SL Mesafesi (%)"
    ]
    header_style(ws, 1, headers, color)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    if state is None:
        ws["A2"] = "Veri bulunamadi."
        return

    positions = state.get("open_positions", {})

    if not positions:
        ws.merge_cells("A2:L2")
        ws["A2"] = "Simdilik acik pozisyon yok."
        ws["A2"].font = Font(italic=True, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")
        return

    for i, (sym, pos) in enumerate(positions.items(), 2):
        side    = pos.get("side", "")
        if isinstance(side, dict):
            side = side.get("value", "")
        entry   = pos.get("entry_price", 0.0)
        current = pos.get("current_price", entry)
        sl      = pos.get("stop_loss", 0.0)
        pnl_u   = pos.get("pnl_usdt", 0.0)
        pnl_p   = pos.get("pnl_pct", 0.0)

        sl_dist = abs(current - sl) / current * 100 if current > 0 and sl > 0 else 0

        row_data = [
            sym, str(side).upper(), entry, current,
            pos.get("amount", 0.0), pos.get("cost_usdt", 0.0),
            sl, pos.get("take_profit", 0.0),
            round(pnl_u, 2), round(pnl_p * 100, 2),
            fmt_dt(pos.get("opened_at")),
            round(sl_dist, 2),
        ]

        is_even = (i % 2 == 0)
        data_row_style(ws, i, len(headers), accent, is_even)

        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if j == 2:
                cell.font = Font(bold=True,
                                  color="005500" if "LONG" in str(side).upper() else "880000")
            if j == 9:
                cell.font = Font(bold=True,
                                  color="005500" if pnl_u >= 0 else "880000")
            if j == 12 and sl_dist < 3.0 and sl_dist > 0:
                cell.font = Font(bold=True, color="CC0000")

    auto_col_width(ws)


def add_signals_sheet(wb: Workbook, ver: dict, signals: List[Dict]) -> None:
    """Sinyal kayitlari sayfasi."""
    ws = wb.create_sheet("Sinyaller")
    color  = ver["color"]
    accent = ver["accent"]

    headers = [
        "Tarih", "Sembol", "TF", "Sinyal Tipi", "Ag. Skor",
        "Guc", "Giris", "SL", "TP", "RR",
        "Guven", "Paper?", "Red Nedeni"
    ]
    header_style(ws, 1, headers, color)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    if not signals:
        ws["A2"] = "Sinyal kaydi bulunamadi."
        ws["A2"].font = Font(italic=True, color="666666")
        return

    # Son 500 sinyal (en yeni uste)
    recent = list(reversed(signals[-500:]))

    for i, sig in enumerate(recent, 2):
        stype = sig.get("signal_type", "")
        rej   = "; ".join(sig.get("rejection_reasons", []))

        row_data = [
            fmt_dt(sig.get("generated_at")),
            sig.get("symbol", ""),
            sig.get("timeframe", ""),
            stype,
            round(sig.get("weighted_score", 0.0), 4),
            sig.get("signal_strength", ""),
            sig.get("entry_price", 0.0),
            sig.get("stop_loss", 0.0),
            sig.get("take_profit", 0.0),
            sig.get("risk_reward_ratio", 0.0),
            sig.get("confidence", 0.0),
            "Evet" if sig.get("is_paper_trade") else "Hayir",
            rej[:120] if rej else "",
        ]

        is_even = (i % 2 == 0)
        data_row_style(ws, i, len(headers), accent, is_even)

        type_colors = {
            "BUY":       "005500",
            "SELL":      "880000",
            "HOLD":      "664400",
            "NO_SIGNAL": "888888",
        }
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if j == 4:
                cell.font = Font(bold=True, color=type_colors.get(stype, "000000"))

    auto_col_width(ws)


def add_trade_history_sheet(wb: Workbook, ver: dict, state: Optional[Dict]) -> None:
    """
    ISLEM GECMISI sayfasi — her acilis ve kapanis emrini net sekilde gosterir.
    Kar/Zarar sutunlari renkli, sebep sutunu aciklayici.
    """
    ws = wb.create_sheet("Islem Gecmisi")
    color  = ver["color"]
    accent = ver["accent"]

    # Sayfa basligi
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"{ver['title'].upper()} — ISLEM GECMISI"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor=color)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Tarih",
        "Sembol",
        "Yon",
        "Islem Tipi",   # ACILIS / KAPANIS
        "Emir Tipi",    # MARKET / LIMIT
        "Fiyat ($)",
        "Miktar",
        "Hacim (USDT)",
        "Komisyon ($)",
        "SL ($)",
        "TP ($)",
        "PnL ($)",
        "PnL (%)",
        "Kapanis Nedeni",
    ]
    header_style(ws, 2, headers, color)
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    if state is None:
        ws.merge_cells("A3:N3")
        ws["A3"] = "Bu versiyon icin portfolyo verisi bulunamadi."
        ws["A3"].font = Font(color="FF0000", bold=True)
        ws["A3"].alignment = Alignment(horizontal="center")
        return

    orders = state.get("orders", [])

    # FILLED emirleri al, tarihe gore sirala (eskiden yeniye)
    filled = [o for o in orders if o.get("status") == "filled"]
    try:
        filled = sorted(filled, key=lambda o: o.get("timestamp", ""))
    except Exception:
        pass

    if not filled:
        ws.merge_cells("A3:N3")
        ws["A3"] = "Henuz hicbir islem gerceklesmedi."
        ws["A3"].font = Font(italic=True, color="666666")
        ws["A3"].alignment = Alignment(horizontal="center")
        auto_col_width(ws)
        return

    # Toplam istatistik sayaclar
    total_pnl   = 0.0
    total_fees  = 0.0
    win_count   = 0
    loss_count  = 0
    trade_count = 0

    for i, o in enumerate(filled, 3):
        side       = o.get("side", "").upper()
        close_rsn  = o.get("close_reason")
        pnl_u      = o.get("pnl_usdt", 0.0)
        pnl_p      = o.get("pnl_pct", 0.0)
        price      = o.get("price", 0.0)
        amount     = o.get("amount", 0.0)
        fee        = o.get("fee_usdt", 0.0)
        o_type     = o.get("order_type", "market").upper()

        # Acilis mi kapanis mi?
        is_close = bool(close_rsn)
        trade_type = "KAPANIS" if is_close else "ACILIS"

        # Hacim
        hacim = round(price * amount, 2)

        row_data = [
            fmt_dt(o.get("timestamp")),
            o.get("symbol", ""),
            side,
            trade_type,
            o_type,
            price,
            round(amount, 6),
            hacim,
            round(fee, 4),
            o.get("stop_loss", 0.0),
            o.get("take_profit", 0.0),
            round(pnl_u, 2) if is_close else "",
            f"{pnl_p * 100:+.2f}%" if is_close and pnl_p != 0 else "",
            close_rsn if close_rsn else "—",
        ]

        is_even = (i % 2 == 0)
        # Kapanis emirleri icin arka plan rengini PnL'e gore ayarla
        if is_close:
            if pnl_u >= 0:
                row_bg = "D6FFDA"  # Acik yesil
            else:
                row_bg = "FFD6D6"  # Acik kirmizi
            fill = PatternFill(fill_type="solid", fgColor=row_bg)
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = fill
                ws.cell(row=i, column=col).border = make_border()
                ws.cell(row=i, column=col).alignment = Alignment(vertical="center")
        else:
            data_row_style(ws, i, len(headers), accent, is_even)

        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)

            # Yon rengi
            if j == 3:
                cell.font = Font(bold=True,
                                  color="005500" if side == "BUY" else "880000")
            # Islem tipi rengi
            if j == 4:
                cell.font = Font(bold=True,
                                  color="0070C0" if trade_type == "ACILIS" else "7030A0")
            # PnL rengi
            if j == 12 and is_close and val != "":
                cell.font = Font(bold=True,
                                  color="005500" if pnl_u >= 0 else "CC0000")
            # Kapanis nedeni rengi
            if j == 14 and is_close:
                reason_colors = {
                    "TAKE PROFIT TRIGGERED": "005500",
                    "STOP LOSS TRIGGERED":   "CC0000",
                    "TRAILING STOP TRIGGERED": "885500",
                    "STRATEGY EXIT":          "0070C0",
                    "TIME-BASED EXIT TRIGGERED": "7030A0",
                    "LIQUIDATION":            "CC0000",
                }
                cell.font = Font(bold=True,
                                  color=reason_colors.get(str(val), "333333"))

        # Istatistik guncelle
        if is_close:
            trade_count += 1
            total_pnl   += pnl_u
            total_fees  += fee
            if pnl_u >= 0:
                win_count += 1
            else:
                loss_count += 1

    # ── OZET SATIRI ──────────────────────────────────────────────────────────
    last_row = 3 + len(filled)
    ws.row_dimensions[last_row].height = 20

    ws.merge_cells(f"A{last_row}:K{last_row}")
    sum_cell = ws[f"A{last_row}"]
    win_rate = win_count / trade_count * 100 if trade_count > 0 else 0
    sum_cell.value = (
        f"TOPLAM: {trade_count} kapali islem | "
        f"Kazanc: {win_count} | Kayip: {loss_count} | "
        f"Basari: %{win_rate:.1f} | "
        f"Toplam Komisyon: ${total_fees:.2f}"
    )
    sum_cell.font = Font(bold=True, size=10, color="FFFFFF")
    sum_cell.fill = PatternFill(fill_type="solid", fgColor=color)
    sum_cell.alignment = Alignment(horizontal="left", indent=1)

    pnl_cell = ws.cell(row=last_row, column=12)
    pnl_cell.value = round(total_pnl, 2)
    pnl_cell.font = Font(bold=True, size=11,
                          color="005500" if total_pnl >= 0 else "CC0000")
    pnl_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFC0")
    pnl_cell.alignment = Alignment(horizontal="right")
    pnl_cell.border = make_border(thin=False)

    label_cell = ws.cell(row=last_row, column=13)
    label_cell.value = f"{'KAR' if total_pnl >= 0 else 'ZARAR'} ← NET PnL"
    label_cell.font = Font(bold=True, color="005500" if total_pnl >= 0 else "CC0000")
    label_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFC0")

    auto_col_width(ws)




def add_stats_sheet(wb: Workbook, ver: dict, state: Optional[Dict], signals: List[Dict]) -> None:
    """Sembol / timeframe bazinda performans ozeti."""
    ws = wb.create_sheet("Sembol Analizi")
    color  = ver["color"]
    accent = ver["accent"]

    ws.merge_cells("A1:G1")
    ws["A1"] = "SEMBOL BAZLI PERFORMANS"
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=color)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    if state is None:
        ws["A3"] = "Veri yok."
        return

    orders = [o for o in state.get("orders", []) if o.get("close_reason")]

    # Sembol bazinda istatistik
    sym_stats: Dict[str, Dict] = {}
    for o in orders:
        sym  = o.get("symbol", "UNKNOWN")
        pnl  = o.get("pnl_usdt", 0.0)
        won  = pnl >= 0
        if sym not in sym_stats:
            sym_stats[sym] = {"trades": 0, "wins": 0, "pnl": 0.0, "fees": 0.0}
        sym_stats[sym]["trades"] += 1
        sym_stats[sym]["wins"]   += 1 if won else 0
        sym_stats[sym]["pnl"]    += pnl
        sym_stats[sym]["fees"]   += o.get("fee_usdt", 0.0)

    headers = ["Sembol", "Islem Sayisi", "Kazanc", "Kayip", "Basari %", "Net PnL", "Komisyon"]
    header_style(ws, 3, headers, color)
    ws.row_dimensions[3].height = 18

    row = 4
    for sym, s in sorted(sym_stats.items(), key=lambda x: x[1]["pnl"], reverse=True):
        wins   = s["wins"]
        trades = s["trades"]
        losses = trades - wins
        wr     = wins / trades * 100 if trades > 0 else 0
        pnl    = s["pnl"]

        is_even = (row % 2 == 0)
        data_row_style(ws, row, len(headers), accent, is_even)

        row_data = [sym, trades, wins, losses, round(wr, 1), round(pnl, 2), round(s["fees"], 2)]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=j, value=val)
            if j == 5:
                cell.font = Font(bold=True, color="005500" if wr >= 50 else "880000")
            if j == 6:
                cell.font = Font(bold=True, color="005500" if pnl >= 0 else "880000")

        row += 1

    if not sym_stats:
        ws["A4"] = "Henuz kapali islem yok."

    # Sinyal tipi dagilimi
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "SINYAL TIP DAGILIMI"
    ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f"A{row}"].fill = PatternFill(fill_type="solid", fgColor=color)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 20
    row += 1

    sig_headers = ["Sinyal Tipi", "Sayi", "Oran (%)"]
    header_style(ws, row, sig_headers, color)
    row += 1

    from collections import Counter
    sig_counts = Counter(s.get("signal_type", "?") for s in signals)
    total_sigs = len(signals)
    for stype, cnt in sig_counts.most_common():
        is_even = (row % 2 == 0)
        data_row_style(ws, row, 3, accent, is_even)
        ws.cell(row=row, column=1, value=stype)
        ws.cell(row=row, column=2, value=cnt)
        pct_cell = ws.cell(row=row, column=3, value=round(cnt / total_sigs * 100, 1) if total_sigs else 0)
        row += 1

    auto_col_width(ws)


# ---- Ana Rapor Olusturma ----

def generate_report(ver: dict, output_dir: Path) -> Path:
    """Tek bir versiyon icin Excel raporu olusturur."""
    label   = ver["label"]
    log_dir = ver["log_dir"]

    print(f"\n[{label}] Rapor olusturuluyor...")
    print(f"  Log dizini: {log_dir}")

    state   = load_portfolio(log_dir)
    signals = load_signals(log_dir)

    print(f"  Portfolyo durumu: {'OK' if state else 'BULUNAMADI'}")
    print(f"  Sinyal kaydi    : {len(signals)} adet")

    wb = Workbook()

    add_summary_sheet(wb, ver, state)
    add_trade_history_sheet(wb, ver, state)
    add_orders_sheet(wb, ver, state)
    add_positions_sheet(wb, ver, state)
    add_signals_sheet(wb, ver, signals)
    add_stats_sheet(wb, ver, state, signals)

    now_str  = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"rapor_{label.replace('.', '_')}_{now_str}.xlsx"
    out_path = output_dir / filename

    wb.save(out_path)
    print(f"  Kaydedildi: {out_path}")
    return out_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="ANTIGRAVITI -- Her versiyon icin Excel raporu olusturur"
    )
    parser.add_argument(
        "--output-dir", type=str, default="raporlar",
        help="Raporlarin kaydedilecegi klasor (varsayilan: raporlar/)"
    )
    parser.add_argument(
        "--version", type=str, default=None, choices=["v1", "v2", "v2.1"],
        help="Sadece belirli bir versiyon icin rapor uret"
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    output_dir = PROJECT_ROOT / args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("[*] ANTIGRAVITI -- MULTI-VERSION EXCEL RAPOR OLUSTURUCU")
    print("=" * 60)
    print(f"Cikti dizini: {output_dir}")

    versions_to_run = VERSIONS
    if args.version:
        versions_to_run = [v for v in VERSIONS if v["label"].lower() == args.version]

    generated = []
    for ver in versions_to_run:
        try:
            path = generate_report(ver, output_dir)
            generated.append((ver["label"], path))
        except Exception as e:
            print(f"  HATA [{ver['label']}]: {e}")

    print()
    print("=" * 60)
    print("[OK] Raporlar olusturuldu:")
    for label, path in generated:
        print(f"  {label:6s} -> {path.name}")
    print("=" * 60)


if __name__ == "__main__":
    main()
